"""
AUREVIX — Unit Test Suite: Production-Grade Real-Time BI Export Briefing
Validates dynamic KPI computation, metadata profiling, temporal trend breakdowns,
profitability analysis, segment/geographic analysis, customer concentration (Pareto),
data quality scorecards, 10-sheet Excel workbooks, pure-python PDF generation,
formula injection sanitization, and security audit logging.
"""
import io
import pytest
import openpyxl
import pandas as pd
from pathlib import Path

from dashboard.analytics.schema_detector import SchemaDetector
from dashboard.analytics.profiler import DataProfiler
from dashboard.analytics.metric_engine import MetricEngine
from dashboard.analytics.insight_engine import InsightEngine
from dashboard.analytics.anomaly_engine import AnomalyEngine
from dashboard.analytics.report_generator import ExecutiveReportGenerator
from dashboard.analytics.pdf_generator import AUREVIXPDFGenerator, PDFDocument
from dashboard.analytics.security_utils import sanitize_for_spreadsheet_export
from dashboard.analytics.security_audit import SecurityAuditLogger, SecurityEventType


@pytest.fixture
def business_df():
    csv_path = Path("data/samples/aurevix_business_test_dataset.csv")
    assert csv_path.exists(), f"Sample business dataset not found at {csv_path}"
    return pd.read_csv(csv_path)


@pytest.fixture
def business_res(business_df):
    schema = SchemaDetector.detect_schema(business_df)
    prof = DataProfiler.profile(business_df, schema)
    kpis = MetricEngine.calculate_metrics(business_df, schema)
    insights = InsightEngine.generate_insights(business_df, schema, kpis)
    anomalies = AnomalyEngine.detect_anomalies(business_df, schema, kpis)
    return {
        "dataset_id": "test_dataset_aurevix_hash_001",
        "dataset_name": "aurevix_business_test_dataset.csv",
        "schema": schema,
        "profile": prof,
        "kpis": kpis,
        "insights": insights,
        "anomalies": anomalies,
        "analysis_time_ms": 14.5
    }


def test_markdown_briefing_production_grade_headers_and_metadata(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "AUREVIX — Executive Business Intelligence Report" in rep
    assert "Executive Intelligence Briefing" in rep
    assert "aurevix_business_test_dataset.csv" in rep
    assert "Domain Classification" in rep
    assert "Reporting Period" in rep
    assert "COMPLETED (VERIFIED)" in rep
    assert "Records Quarantined" in rep
    assert "Records Rejected" in rep
    assert "Report Correlation ID" in rep


def test_dynamic_kpis_calculation_on_business_dataset(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    # Total Revenue: $1,719,803.88
    assert "1,719,803.88" in rep
    # Total Transactions: 36
    assert "36" in rep
    # Total Quantity: 21,758
    assert "21,758" in rep
    # Average Transaction Value: $47,772.33
    assert "47,772.33" in rep
    # Total Net Profit: $426,879.92
    assert "426,879.92" in rep
    # Profit Margin: 24.8%
    assert "24.8%" in rep
    # Unique Accounts: 15
    assert "15" in rep
    # Data Quality Score: 100.0%
    assert "100.0%" in rep


def test_revenue_trend_and_fallback(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Chronological Revenue Trend Breakdown" in rep
    assert "Period" in rep
    assert "% Volume" in rep

    # Dataset without date column
    df_nodate = business_df.drop(columns=["Date"])
    schema_nodate = SchemaDetector.detect_schema(df_nodate)
    kpis_nodate = MetricEngine.calculate_metrics(df_nodate, schema_nodate)
    res_nodate = {
        "dataset_name": "no_date.csv",
        "schema": schema_nodate,
        "profile": business_res["profile"],
        "kpis": kpis_nodate,
        "insights": [],
        "anomalies": []
    }
    rep_nodate = ExecutiveReportGenerator.generate_report(res_nodate, df_nodate)
    assert "Revenue trend unavailable for this dataset because no valid temporal comparison field was detected." in rep_nodate


def test_profitability_analysis_and_fallback(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Profitability & Margin Analysis" in rep
    assert "Total Net Profit" in rep
    assert "Net Operating Profit Margin" in rep
    assert "Average Profit per Transaction" in rep
    assert "Highest-Profit Segment" in rep

    # Dataset without profit
    df_noprofit = business_df.drop(columns=["Profit"])
    schema_noprofit = SchemaDetector.detect_schema(df_noprofit)
    kpis_noprofit = MetricEngine.calculate_metrics(df_noprofit, schema_noprofit)
    res_noprofit = {
        "dataset_name": "no_profit.csv",
        "schema": schema_noprofit,
        "profile": business_res["profile"],
        "kpis": kpis_noprofit,
        "insights": [],
        "anomalies": []
    }
    rep_noprofit = ExecutiveReportGenerator.generate_report(res_noprofit, df_noprofit)
    assert "Profitability metrics unavailable — no profit or cost dimension detected in dataset." in rep_noprofit


def test_segment_analysis_and_fallback(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Segment Performance (Categories & Products)" in rep
    assert "TOP PERFORMING SEGMENT" in rep
    assert "Electronics" in rep

    # Dataset without category or any categorical dimension
    df_nocat = business_df.drop(columns=["Category", "Region"])
    schema_nocat = SchemaDetector.detect_schema(df_nocat)
    kpis_nocat = MetricEngine.calculate_metrics(df_nocat, schema_nocat)
    res_nocat = {
        "dataset_name": "no_cat.csv",
        "schema": schema_nocat,
        "profile": business_res["profile"],
        "kpis": kpis_nocat,
        "insights": [],
        "anomalies": []
    }
    rep_nocat = ExecutiveReportGenerator.generate_report(res_nocat, df_nocat)
    assert "Segment analysis unavailable" in rep_nocat


def test_geographic_analysis_and_fallback(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Geographic Performance" in rep
    assert "Top Regional Market" in rep
    assert "North" in rep

    # Dataset without region
    df_noreg = business_df.drop(columns=["Region"])
    schema_noreg = SchemaDetector.detect_schema(df_noreg)
    kpis_noreg = MetricEngine.calculate_metrics(df_noreg, schema_noreg)
    res_noreg = {
        "dataset_name": "no_reg.csv",
        "schema": schema_noreg,
        "profile": business_res["profile"],
        "kpis": kpis_noreg,
        "insights": [],
        "anomalies": []
    }
    rep_noreg = ExecutiveReportGenerator.generate_report(res_noreg, df_noreg)
    assert "Geographic analysis unavailable for this dataset." in rep_noreg


def test_customer_concentration_and_pareto(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Customer & Account Concentration" in rep
    assert "Top 20% of accounts contribute" in rep
    assert "26.0% of total revenue" in rep
    assert "Top 5 Accounts Share:" in rep
    assert "Top 10 Accounts Share:" in rep

    # Dataset without customer
    df_nocust = business_df.drop(columns=["Customer_ID"])
    schema_nocust = SchemaDetector.detect_schema(df_nocust)
    kpis_nocust = MetricEngine.calculate_metrics(df_nocust, schema_nocust)
    res_nocust = {
        "dataset_name": "no_cust.csv",
        "schema": schema_nocust,
        "profile": business_res["profile"],
        "kpis": kpis_nocust,
        "insights": [],
        "anomalies": []
    }
    rep_nocust = ExecutiveReportGenerator.generate_report(res_nocust, df_nocust)
    assert "Customer & Account Concentration analysis unavailable — no customer/account identifier detected." in rep_nocust


def test_data_quality_scorecard_and_pillars(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Data Quality Scorecard" in rep
    assert "Completeness Pillar" in rep
    assert "Validity Pillar" in rep
    assert "Consistency Pillar" in rep
    assert "Uniqueness Pillar" in rep
    assert "100.0%" in rep


def test_automated_business_insights_and_alerts(business_df, business_res):
    rep = ExecutiveReportGenerator.generate_report(business_res, business_df)
    assert "Automated Business Insights" in rep
    assert "Observation:" in rep
    assert "Underlying Driver:" in rep
    assert "Strategic Impact:" in rep
    assert "Executive Alerts & Exceptions" in rep
    assert "✓ No material data-quality or analytical exceptions detected." in rep


def test_excel_10_sheets_generation(business_df, business_res):
    excel_bytes = ExecutiveReportGenerator.generate_excel_report(business_res, business_df)
    assert len(excel_bytes) > 5000

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    expected_sheets = [
        "Executive Summary", "KPI Summary", "Revenue Analysis", "Segment Analysis",
        "Geographic Analysis", "Customer Analysis", "Data Quality", "Business Insights",
        "Audit Metadata", "Cleaned Data"
    ]
    assert wb.sheetnames == expected_sheets

    # Verify Executive Summary
    ws_exec = wb["Executive Summary"]
    assert ws_exec["A1"].value == "AUREVIX  |  Executive Intelligence Briefing"
    assert ws_exec.column_dimensions["A"].width >= 14

    # Verify KPI Summary
    ws_kpi = wb["KPI Summary"]
    assert ws_kpi["A4"].value == "KPI ID"

    # Verify Cleaned Data has 37 rows (1 header + 36 data)
    ws_data = wb["Cleaned Data"]
    assert ws_data.max_row == 37


def test_pdf_report_generation(business_df, business_res):
    pdf_bytes = ExecutiveReportGenerator.generate_pdf_report(business_res, business_df)
    assert pdf_bytes.startswith(b"%PDF-1.4")
    assert b"%%EOF" in pdf_bytes
    assert len(pdf_bytes) > 2000


def test_formula_injection_sanitization_in_exports(business_df):
    df_danger = business_df.copy()
    df_danger.loc[0, "Customer_ID"] = "=CMD|' /C calc'!A0"
    df_danger.loc[1, "Customer_ID"] = "+cmd|' /C notepad'!A0"
    df_danger.loc[2, "Customer_ID"] = "-danger_macro"
    df_danger.loc[3, "Customer_ID"] = "@SUM(A1:A10)"
    df_danger.loc[4, "Customer_ID"] = "   =EVIL_FUNCTION()"

    sanitized = sanitize_for_spreadsheet_export(df_danger)
    for idx in range(5):
        val = str(sanitized.loc[idx, "Customer_ID"])
        assert val.startswith("'"), f"Formula not sanitized at row {idx}: {val}"


def test_security_audit_logging_on_export(business_res):
    event = SecurityAuditLogger.log_event(
        event_type=SecurityEventType.DATA_EXPORT,
        source="dashboard.workspace.export",
        dataset_id=business_res["dataset_id"],
        metadata={"export_format": "xlsx", "dataset_name": business_res["dataset_name"]}
    )
    assert event["event_type"] == SecurityEventType.DATA_EXPORT
    assert event["outcome"] == "SUCCESS"
    assert "event_hash" in event
    assert "previous_event_hash" in event


def test_render_export_center_all_formats_including_parquet(business_df, business_res, monkeypatch):
    import importlib
    import streamlit as st
    dw = importlib.import_module("dashboard.pages.10_Data_Workspace")
    assert hasattr(dw, "io"), "10_Data_Workspace module must have io imported"

    rendered_buttons = []
    def mock_download_button(label, data, file_name, mime, on_click=None, args=None, use_container_width=True):
        rendered_buttons.append({
            "label": label,
            "file_name": file_name,
            "data_len": len(data) if data is not None else 0,
            "mime": mime
        })

    monkeypatch.setattr(st, "download_button", mock_download_button)
    dw.render_export_center(business_df, business_res, user_active=True)

    assert len(rendered_buttons) == 7
    parquet_btn = next(b for b in rendered_buttons if b["file_name"].endswith(".parquet"))
    assert parquet_btn["data_len"] > 0
    assert parquet_btn["mime"] == "application/octet-stream"

    excel_btn = next(b for b in rendered_buttons if b["file_name"].endswith(".xlsx"))
    assert excel_btn["data_len"] > 0

    csv_btn = next(b for b in rendered_buttons if b["file_name"].endswith(".csv"))
    assert csv_btn["data_len"] > 0

    pdf_btn = next(b for b in rendered_buttons if b["file_name"].endswith(".pdf"))
    assert pdf_btn["data_len"] > 0

